import openpyxl as xl
from openxl.chart import BarChart, Reference


wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet.cell(1, 1)
number_row = sheet.max_row
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value() * 0.95
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


values = Reference(sheet, min_value = 2, 
                    max_value = sheet.max_row, 
                    max_col = 4, min_col = 4)

chart = Barchart()
chart.add_data(values)
sheet.add_chart(chart, "e2")

wb.save("transaction2.xlsx")